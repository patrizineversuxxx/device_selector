import json
import openpyxl
from Data import *


def get_data_from_xlsx(spreadsheet, department_map, user_map, device_list):
    for row in spreadsheet.iter_rows(2):

        device_name = row[0].value
        device_id = row[1].value
        device_group = row[2].value
        device_os = row[3].value
        device_last_checkin_date = row[4].value

        user_id = row[5].value
        user_name = row[6].value
        user_mail = row[7].value
        user_manager_name = row[8].value
        user_manager_mail = row[9].value
        user_job_title = row[10].value
        user_location = row[11].value

        department_name = row[12].value
        department_cost_center = row[13].value

        device = Device(id=device_id, name=device_name, group=device_group,
                        os=device_os, last_checkin_date=device_last_checkin_date)
        device_list.append(device)

        if user_name in user_map:
            user_map[user_name].add_device(device)

        else:
            user = User(id=user_id, name=user_name, mail=user_mail, manager_name=user_manager_name, manager_mail=user_manager_mail,
                        job_title=user_job_title, location=user_location, device_list=[device])
            user_map[user.name] = user

            if department_name in department_map:
                department_map[department_name].add_user(user)

            else:
                department = Department(
                    name=department_name, cost_center=department_cost_center, user_list=[user])
                department_map[department_name] = department


def open_json(path: str):
    f = open(path)
    records = json.load(f)
    f.close()
    return records


def get_data_from_json(params: dict):
    records = open_json(params['path_user'])

    department_map = {}
    user_map = {}

    for record in records:
        user_id = record['id']

        user = User(id=user_id, name=record['displayName'], mail=record['mail'], manager_name=record['manager_name'], manager_mail=record['manager_mail'],
                    job_title=record['jobTitle'], location=record['officeLocation'], device_list=[])

        user_map[user_id] = user

        department_name = record['department']
        
        if department_name in department_map:
            department_map[department_name].user_list.append(user)
        else:
            department_map[department_name] = Department(
                name=department_name, cost_center=0, user_list=[])

    records = open_json(params['path_device'])

    for record in records:
        if record['usersLoggedOn'] == []:
            continue
        else:
            device_last_checkin_date = record['usersLoggedOn'][0]['lastLogOnDateTime']

        device_enrollment_type = record['deviceEnrollmentType']
        device_os = record['operatingSystem']
        device_group = ""

        if (device_enrollment_type == "windowsAzureADJoin") & (device_os == "Windows"):
            device_group = "AAD_Joined"
        else:
            if (device_enrollment_type == "windowsCoManagement") & (device_os == "Windows"):
                device_group = "Hybrid_Joined"
            else:
                if (device_enrollment_type == "userEnrollment") & (device_os == "macOS"):
                    device_group = "macOS"

        device = Device(id=record['azureActiveDirectoryDeviceId'], name=record['deviceName'], group=device_group,
                        os=device_os, last_checkin_date=device_last_checkin_date)

        device_user_id = record['userId']

        if device_user_id in user_map:
            user_map[device_user_id].device_list.append(device)
        else:
            continue

    return department_map


def save_data_to_xlsx_prepational_step(result_map):
    result_book = openpyxl.Workbook()
    result_sheet = result_book.active

    result_sheet.cell(row=1, column=1, value='Device_name')
    result_sheet.cell(row=1, column=2, value='Device_id')
    result_sheet.cell(row=1, column=3, value='Device_Group')
    result_sheet.cell(row=1, column=4, value='Device_os')
    result_sheet.cell(row=1, column=5, value='Last_checkin_date')
    result_sheet.cell(row=1, column=6, value='User_id')
    result_sheet.cell(row=1, column=7, value='User_name')
    result_sheet.cell(row=1, column=8, value='User_mail')
    result_sheet.cell(row=1, column=9, value='Manager_name')
    result_sheet.cell(row=1, column=10, value='Manager_mail')
    result_sheet.cell(row=1, column=11, value='Job_title')
    result_sheet.cell(row=1, column=12, value='Location')
    result_sheet.cell(row=1, column=13, value='Department name')
    result_sheet.cell(row=1, column=14, value='Cost center')

    row_counter = 2

    for department in result_map.values():

        department_name = department.name
        cost_center = department.cost_center

        # Iterate over the set of user-device tuples
        for user in department.user_list:
            for device in user.device_list:
                device_name = device.name
                device_id = device.id
                group = device.group
                os = device.os
                last_checkin_date = device.last_checkin_date

                user_id = user.id
                username = user.name
                mail = user.mail
                manager_name = user.manager_name
                manager_mail = user.manager_mail
                job_title = user.job_title
                location = user.location

                result_sheet.cell(row=row_counter, column=1, value=device_name)
                result_sheet.cell(row=row_counter, column=2, value=device_id)
                result_sheet.cell(row=row_counter, column=3, value=group)
                result_sheet.cell(row=row_counter, column=4, value=os)
                result_sheet.cell(row=row_counter, column=5,
                                  value=last_checkin_date)
                result_sheet.cell(row=row_counter, column=6, value=user_id)
                result_sheet.cell(row=row_counter, column=7, value=username)
                result_sheet.cell(row=row_counter, column=8, value=mail)
                result_sheet.cell(row=row_counter, column=9,
                                  value=manager_name)
                result_sheet.cell(row=row_counter, column=10,
                                  value=manager_mail)
                result_sheet.cell(row=row_counter, column=11, value=job_title)
                result_sheet.cell(row=row_counter, column=12, value=location)
                result_sheet.cell(row=row_counter, column=13,
                                  value=department_name)
                result_sheet.cell(row=row_counter, column=14,
                                  value=cost_center)

                row_counter += 1
    result_book.save(r'C:\KEK\KEK.xlsx')


def save_data_to_xlsx(result_map):
    result_book = openpyxl.Workbook()
    result_sheet = result_book.active

    result_sheet.cell(row=1, column=1, value='Device_name')
    result_sheet.cell(row=1, column=2, value='Device_id')
    result_sheet.cell(row=1, column=3, value='Device_Group')
    result_sheet.cell(row=1, column=4, value='Device_os')
    result_sheet.cell(row=1, column=5, value='Last_checkin_date')
    result_sheet.cell(row=1, column=6, value='User_id')
    result_sheet.cell(row=1, column=7, value='User_name')
    result_sheet.cell(row=1, column=8, value='User_mail')
    result_sheet.cell(row=1, column=9, value='Manager_name')
    result_sheet.cell(row=1, column=10, value='Manager_name')
    result_sheet.cell(row=1, column=11, value='Job_title')
    result_sheet.cell(row=1, column=12, value='Location')
    result_sheet.cell(row=1, column=13, value='Department name')
    result_sheet.cell(row=1, column=14, value='Cost center')

    row_counter = 2

    for department, user_device in result_map.items():

        department_name = department.name
        cost_center = department.cost_center

        # Iterate over the set of user-device tuples

        device = list(user_device.values())[0]
        device_name = device.name
        device_id = device.id
        group = device.group
        os = device.os
        last_checkin_date = device.last_checkin_date

        user = list(user_device.keys())[0]
        user_id = user.id
        username = user.name
        mail = user.mail
        manager_name = user.manager_name
        manager_mail = user.manager_mail
        job_title = user.job_title
        location = user.location

        result_sheet.cell(row=row_counter, column=1, value=device_name)
        result_sheet.cell(row=row_counter, column=2, value=device_id)
        result_sheet.cell(row=row_counter, column=3, value=group)
        result_sheet.cell(row=row_counter, column=4, value=os)
        result_sheet.cell(row=row_counter, column=5,
                          value=last_checkin_date)
        result_sheet.cell(row=row_counter, column=6, value=user_id)
        result_sheet.cell(row=row_counter, column=7, value=username)
        result_sheet.cell(row=row_counter, column=8, value=mail)
        result_sheet.cell(row=row_counter, column=9,
                          value=manager_name)
        result_sheet.cell(row=row_counter, column=10,
                          value=manager_mail)
        result_sheet.cell(row=row_counter, column=11, value=job_title)
        result_sheet.cell(row=row_counter, column=12, value=location)
        result_sheet.cell(row=row_counter, column=13,
                          value=department_name)
        result_sheet.cell(row=row_counter, column=14,
                          value=cost_center)

        row_counter += 1
    result_book.save(r'C:\KEK\test.xlsx')
