import openpyxl
from collections import namedtuple
from model.entities import *


# Named tuple for storing table's header
Header = namedtuple(
    'Header', ['Device_name', 'Device_id', 'Device_Group', 'Device_os', 'Device_enrollment_type',
               'Device_type', 'Last_checkin_date', 'User_id', 'User_name', 'User_mail', 'Manager_name',
               'Manager_mail', 'Job_title', 'Location', 'Cost_center', 'Department_name',
               'Current_Device_Pilot_Group', 'Current_User_Pilot_Group']
)
# Named tuple for storing order for rows
Row = namedtuple(
    'Row', ['device', 'user', 'department_name']
)

# Constant defintions for table rows
DEVICE_NAME_IDX = 0
DEVICE_ID_IDX = 1
DEVICE_GROUP_IDX = 2
DEVICE_OS_IDX = 3
DEVICE_ENROLLMENT_TYPE_IDX = 4
DEVICE_TYPE_IDX = 5
DEVICE_LAST_CHECKIN_DATE_IDX = 6
USER_ID_IDX = 7
USER_NAME_IDX = 8
USER_MAIL_IDX = 9
USER_MANAGER_NAME_IDX = 10
USER_MANAGER_MAIL_IDX = 11
USER_JOB_TITLE_IDX = 12
USER_LOCATION_IDX = 13
USER_COST_CENTER_IDX = 14
DEPARTMENT_NAME_IDX = 15
DEVICE_AFFECTED_IDX = 16
USER_AFFECTED_IDX = 17

HEADER = Header._fields


def create_table_header(result_sheet):
    """Create the header row in the result spreadsheet.

    Args:
        result_sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to add the header row to.
    """
    for idx, field in enumerate(HEADER, start=1):
        result_sheet.cell(row=1, column=idx, value=field)


def create_table_row(result_sheet, row_counter, row):
    """Create a row of data in the result spreadsheet.

    Args:
        result_sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to add the row to.
        row_counter (int): The row number to add the data to.
        row (Row): A Row namedtuple containing device, user, and department information.
    """
    device = row.device
    user = row.user
    department_name = row.department_name

    row_data = [
        device.name, device.id, device.group, device.os, device.enrollment_type, device.type,
        device.last_checkin_date, user.id, user.name, user.mail, user.manager_name,
        user.manager_mail, user.job_title, user.location, user.cost_center, department_name,
        device.affected, user.affected
    ]

    for idx, value in enumerate(row_data, start=1):
        result_sheet.cell(row=row_counter, column=idx, value=value)


def extract_user_data(row):
    """Extract user and device data from a spreadsheet row.

    Args:
        row (iterable): A row of data from the spreadsheet.

    Returns:
        Tuple[Device, User, str]: A tuple containing the extracted Device, User, and department name.
    """
    device_name = row[DEVICE_NAME_IDX]
    device_id = row[DEVICE_ID_IDX]
    device_group = row[DEVICE_GROUP_IDX]
    device_os = row[DEVICE_OS_IDX]
    device_enrollment_type = row[DEVICE_ENROLLMENT_TYPE_IDX]
    device_type = row[DEVICE_TYPE_IDX]
    device_last_checkin_date = row[DEVICE_LAST_CHECKIN_DATE_IDX]

    user_id = row[USER_ID_IDX]
    user_name = row[USER_NAME_IDX]
    user_mail = row[USER_MAIL_IDX]
    user_manager_name = row[USER_MANAGER_NAME_IDX]
    user_manager_mail = row[USER_MANAGER_MAIL_IDX]
    user_job_title = row[USER_JOB_TITLE_IDX]
    user_location = row[USER_LOCATION_IDX]
    user_cost_center = row[USER_COST_CENTER_IDX]
    department_name = row[DEPARTMENT_NAME_IDX]
    device_affected = row[DEVICE_AFFECTED_IDX]
    user_affected = row[USER_AFFECTED_IDX]

    device = Device(id=device_id, affected=device_affected, name=device_name, group=device_group,
                    os=device_os, enrollment_type=device_enrollment_type, type=device_type,
                    last_checkin_date=device_last_checkin_date)

    user = User(id=user_id, affected=user_affected, name=user_name, mail=user_mail, manager_name=user_manager_name,
                manager_mail=user_manager_mail, job_title=user_job_title, location=user_location,
                cost_center=user_cost_center)

    return device, user, department_name


def process_spreadsheet(worksheet):
    """Process rows in the spreadsheet and yield extracted user data.

    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to process.

    Yields:
        Tuple[Device, User, str]: A tuple containing the extracted Device, User, and department name.
    """
    for row in worksheet.iter_rows(min_row=2):
        yield extract_user_data(row)


def get_data_from_xlsx(path: str):
    """Extract data from the input spreadsheet.

    Args:
        path (str): The path to the input spreadsheet.

    Returns:
        Tuple[Dict[str, Department], Dict[str, User]]: A tuple containing department and user data dictionaries.
    """
    excel_workbook = openpyxl.load_workbook(path)
    worksheet = excel_workbook.active

    department_map = {}
    user_map = {}

    for device, user, department_name in process_spreadsheet(worksheet):
        if user.name in user_map:
            user_map[user.name].add_device(device)
        else:
            user_map[user.name] = user

            if department_name in department_map:
                department_map[department_name].add_user(user)
            else:
                department = Department(name=department_name, user_list=[user])
                department_map[department_name] = department

    return department_map, user_map


def save_data_to_xlsx(result_map, file_path):
    """Save data to the output spreadsheet.

    Args:
        result_map (Dict[str, Department]): A dictionary containing department data.
        file_path (str): The path to save the output spreadsheet.
    """
    result_book = openpyxl.Workbook()
    result_sheet = result_book.active

    create_table_header(result_sheet)

    row_counter = 2

    for department in result_map.values():
        department_name = department.name

        for user in department.user_list:
            for device in user.device_list:
                row = Row(device=device, user=user,
                          department_name=department_name)
                create_table_row(result_sheet, row_counter, row)
                row_counter += 1

    result_book.save(file_path)
