import openpyxl


def is_vip(job_title):
    # Convert job title to lowercase for case-insensitive matching
    job_title = job_title.lower()

    # Check if the job title contains any VIP keywords
    if "country manager" in job_title or \
       ("sr" in job_title and "manager" in job_title) or \
       ("senior" in job_title and "manager" in job_title) or \
       "gm" in job_title or \
       "president" in job_title or \
       "director" in job_title:
        return True
    else:
        return False


def check_xlsx_for_vip(params):
    # Open the input and output workbooks
    workbook = openpyxl.load_workbook(params['start_file'])
    spreadsheet = workbook.active
    result_book = openpyxl.Workbook()
    result_sheet = result_book.active

    # Iterate over each row in the input workbook
    for row in spreadsheet.iter_rows():
        # Check if the job title is a VIP title
        if is_vip(row[10].value):
            continue  # skip this row

        # If the job title is not a VIP title, copy the row to the output workbook
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        result_sheet.append(row_data)

    # Save the output workbook
    result_book.save(params['middle_file'])
