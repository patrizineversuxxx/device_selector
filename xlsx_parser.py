import openpyxl
from entities import *


def create_table_header(result_sheet):
    result_sheet.cell(row=1, column=1, value='Device_name')
    result_sheet.cell(row=1, column=2, value='Device_id')
    result_sheet.cell(row=1, column=3, value='Device_Group')
    result_sheet.cell(row=1, column=4, value='Device_os')
    result_sheet.cell(row=1, column=5, value='Device_enrollment_type')
    result_sheet.cell(row=1, column=6, value='Device_type')
    result_sheet.cell(row=1, column=7, value='Last_checkin_date')
    result_sheet.cell(row=1, column=8, value='User_id')
    result_sheet.cell(row=1, column=9, value='User_name')
    result_sheet.cell(row=1, column=10, value='User_mail')
    result_sheet.cell(row=1, column=11, value='Manager_name')
    result_sheet.cell(row=1, column=12, value='Manager_mail')
    result_sheet.cell(row=1, column=13, value='Job_title')
    result_sheet.cell(row=1, column=14, value='Location')
    result_sheet.cell(row=1, column=15, value='Cost center')
    result_sheet.cell(row=1, column=16, value='Department name')


def create_table_row(result_sheet, row_counter, device, user, department_name):
    result_sheet.cell(row=row_counter, column=1, value=device.name)
    result_sheet.cell(row=row_counter, column=2, value=device.id)
    result_sheet.cell(row=row_counter, column=3, value=device.group)
    result_sheet.cell(row=row_counter, column=4, value=device.os)
    result_sheet.cell(row=row_counter, column=5, value=device.enrollment_type)
    result_sheet.cell(row=row_counter, column=6, value=device.type)
    result_sheet.cell(row=row_counter, column=7,
                      value=device.last_checkin_date)
    result_sheet.cell(row=row_counter, column=8, value=user.id)
    result_sheet.cell(row=row_counter, column=9, value=user.name)
    result_sheet.cell(row=row_counter, column=10, value=user.mail)
    result_sheet.cell(row=row_counter, column=11,
                      value=user.manager_name)
    result_sheet.cell(row=row_counter, column=12,
                      value=user.manager_mail)
    result_sheet.cell(row=row_counter, column=13, value=user.job_title)
    result_sheet.cell(row=row_counter, column=14, value=user.location)
    result_sheet.cell(row=row_counter, column=15,
                      value=user.cost_center)
    result_sheet.cell(row=row_counter, column=16,
                      value=department_name)


def get_data_from_xlsx(path: str):

    workbook = openpyxl.load_workbook(path)
    spreadsheet = workbook.active

    user_map = {}
    department_map = {}

    for row in spreadsheet.iter_rows(2):

        device_name = row[0].value
        device_id = row[1].value
        device_group = row[2].value
        device_os = row[3].value
        device_enrollment_type = row[4].value
        device_type = row[5].value
        device_last_checkin_date = row[6].value

        user_id = row[7].value
        user_name = row[8].value
        user_mail = row[9].value
        user_manager_name = row[10].value
        user_manager_mail = row[11].value
        user_job_title = row[12].value
        user_location = row[13].value
        user_cost_center = row[14].value
        department_name  = row[15].value
        

        device = Device(id=device_id, name=device_name, group=device_group,
                        os=device_os, enrollment_type=device_enrollment_type, type=device_type,
                        last_checkin_date=device_last_checkin_date)

        if user_name in user_map:
            user_map[user_name].add_device(device)

        else:
            user = User(id=user_id, name=user_name, mail=user_mail, manager_name=user_manager_name, manager_mail=user_manager_mail,
                        job_title=user_job_title, location=user_location, cost_center=user_cost_center, device_list=[device])
            user_map[user.name] = user

            if department_name in department_map:
                department_map[department_name].add_user(user)

            else:
                department = Department(
                    name=department_name, user_list=[user])
                department_map[department_name] = department

    return department_map, user_map


def save_data_to_xlsx_prepational_step(result_map, file_path):
    result_book = openpyxl.Workbook()
    result_sheet = result_book.active

    create_table_header(result_sheet)

    row_counter = 2

    for department in result_map.values():

        department_name = department.name

        # Iterate over the set of user-device tuples
        for user in department.user_list:
            for device in user.device_list:

                create_table_row(result_sheet, row_counter, device, user,
                                 department_name)
                row_counter += 1

    result_book.save(file_path)


def save_data_to_xlsx(result_map, path):
    result_book = openpyxl.Workbook()
    result_sheet = result_book.active

    create_table_header(result_sheet)

    row_counter = 2

    for department, user_device in result_map.items():

        department_name = department.name

        # Iterate over the set of user-device tuples
        device = list(user_device.values())[0]
        user = list(user_device.keys())[0]

        create_table_row(result_sheet, row_counter, device, user,
                         department_name)

        row_counter += 1

    result_book.save(path)
