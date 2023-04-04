import openpyxl


def check_vip(kek):
    kek = kek.lower()
    if (kek.__contains__("сountry") & kek.__contains__("manager")) | (kek.__contains__("sr") & kek.__contains__("manager")) | (kek.__contains__("senior") & kek.__contains__("manager")) | kek.__contains__("gm") | kek.__contains__("president") | kek.__contains__("director"):
        return True
    else:
        return False

def check_xlsx_for_vip(params):
    workbook = openpyxl.load_workbook("C:\KEK\KEK.xlsx")
    spreadsheet = workbook.active
    result_book = openpyxl.Workbook()
    result_sheet = result_book.active

    for row in spreadsheet.iter_rows():
        if check_vip(row[10].value):
            continue
        else:
            row_data = []
            # Iterate through the cells in the current row
            for cell in row:
                # Append the cell value to the row data list
                row_data.append(cell.value)
            # Write the row data to the destination worksheet
            result_sheet.append(row_data)

    result_book.save(r"C:\KEK\vip_checked_kek.xlsx")
